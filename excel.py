#encoding=utf8
import copy

from util import *
from openpyxl import load_workbook, Workbook
reload(sys)
sys.setdefaultencoding('utf8')


class MExcel():
    TYPE_0 = 'singe'
    TYPE_1 = 'many'
    TYPE_1_0 = 'many_to_right'
    TYPE_1_1 = 'many_to_bottom'
    def __init__(self, path):
        self.wb = load_workbook(path,data_only=True)

    def getSheet(self, sheetName):
        return self.wb.get_sheet_by_name(sheetName)

    def isEnd(self, value, mList=endKeys):
        if not value:
            return True
        for key in mList:
            if re.findall(key.decode('utf8'), str(value).decode('utf8')):
                return True
        return False

    def parseArea(self, area):
        areas = area.split(':')
        min_col = re.findall('\D+', areas[0])[0]
        min_row = re.findall('\d+', areas[0])[0]
        max_col = re.findall('\D+', areas[1])[0]
        max_row = re.findall('\d+', areas[1])[0]
        return [min_col,int(min_row),max_col,int(max_row)]

    def parseTitle(self, area, sheet):
        if sheet[area][0][0].value:
            return {
                'title': sheet[area][0][0].value,
                'titleParse': self.parseArea(area),
                'mergedCells': []
            }
        return None

    def inArea(self, small, bigStr):
        small = self.parseArea(small)
        big = self.parseArea(bigStr)
        return big[0] <= small[0] and big[1] <= small[1] and big[2] >= small[2] and big[3] >= small[3], bigStr

    def getWidth(self, table):
        temp = self.parseArea(table['range'])
        return ord(table['son'][0]['zbfm'][0].column)-ord(temp[0])

    def getEndRow(self, table, sheet):
        for row in sheet[table['range']]:
            if row[0].value:
                if self.isEnd(row[0].value):
                    return row[0].row
        return -1

    def similar(self, stra, strb, percent=50):
        stra,strb,count = str(stra),str(strb),0
        stra = list(jieba.cut(stra))
        strb = list(jieba.cut(strb))
        for i in stra:
            if i in strb:
                count+=1
        return count*100/len(stra)>percent and count*100/len(strb)>percent

    def isIgnoreZbfm(self, inStr):
        for i in ignoreZbfms:
            if inStr==i.decode('utf8'):
                return True
        return False

    def dealZbzm(self, zbzms, sheet):
        temp1, temp2 = {},[]
        for zbzm in zbzms:
            if zbzm.value:
                zbzm.value = delPreStr(zbzm.value)
                if zbzm.value in temp1:
                    temp2.append(zbzm)
                    if temp1[zbzm.value] not in temp2:
                        temp2.append(temp1[zbzm.value])
                else:
                    temp1[zbzm.value] = zbzm
        for zbzm in zbzms:
            if zbzm in temp2:
                if re.findall('#', zbzm.value):
                    zbzm.value = sheet[zbzm.column+str(zbzm.row-1)].value+zbzm.value
                else:
                    i,pre = 1,None
                    while not pre:
                        cell = sheet[zbzm.column+str(zbzm.row-i)]
                        if cell.value:
                            if re.findall('#', cell.value):
                                pre = cell
                        i += 1
                    if pre and sheet[zbzm.column+str(pre.row-1)].value:
                        zbzm.value = sheet[zbzm.column+str(pre.row-1)].value+'#'+zbzm.value
                print zbzm.value
        return zbzms

    def findRow(self, table, sheet, min, type=isZbfm):
        if min == 0:
            return []
        for row in sheet[table['range']]:
            count = 0
            for cell in row:
                if cell.value:
                    if type(cell.value):
                        count += 1
            if count >= min:
                return list(row)
        if 'zbfm' not in table:
            return self.findRow(table, sheet, min-1)

    def getObjectByTitle(self, table):
        for key in jieba.cut(table['title']):
            if re.findall(u'[区市]', key):
                return key
        return object_name

    def getZbzmByTitle(self, table):
        title = table['title']
        temp = re.findall(u'[\(\（].+[\）\)]', title)
        if temp:
            title = title.replace(temp[0], '')
        return title

    def getTimeCodeByTitle(self, table):
        temp = re.findall(u'\d{4}年[\d-]+月', table['title'])
        if temp:
            temp2 = re.findall(u'\d+-', temp[0])
            if temp2:
                temp[0] = temp[0].replace(temp2[0], '')
            code = temp[0].replace('年','').replace('月','')
            if len(code)==5:
                code = code[:4]+'0'+code[4:]
            return code[:4]+'年'+code[4:]+'月'
        else:
            return time_code

    def getZbObject(self, table, sheet):
        tempPaser = self.parseArea(table['range'])
        end = self.getEndRow(table, sheet)-1
        end = str(tempPaser[3] if end==-2 else end)
        if table['type'] == self.TYPE_0:
            table['zbfm'] = self.findRow(table, sheet, 2)
            table['firstZbfm'] = table['zbfm'][0]
            zbmc,objectmc,allObject = None,{},[]
            for zbfm in table['zbfm']:
                if zbfm.value:
                    if isZbmc(zbfm.value):
                        zbmc = zbfm
                        break
                    temp = isObject(zbfm.value)
                    if temp:
                        objectmc[temp['text']] = zbfm
                        allObject.append(temp)

            width = 0# 改动只为了加入指标为两列的处理
            mc = zbmc if zbmc else objectmc[sorted(allObject, key=lambda o:o['priority']).pop()['text']]
            mcStr = 'zbzm' if zbmc else 'object'
            if not sheet[chr(ord(mc.column)+1)+str(mc.row)].value:
                width += 1

            # 处理 同比（±%）、同比（消除物价因素±%）等情况
            repeater1,repeater2 = {},[]
            for zbfm in table['zbfm']:
                if zbfm.value:
                    tempValue = isZbfm(zbfm.value, getStr=True)
                    if tempValue not in repeater1:
                        repeater1[tempValue] = zbfm
                    else:
                        if repeater1[tempValue] not in repeater2:
                            repeater2.append(repeater1[tempValue])
                        repeater2.append(zbfm)
            for zbfm in table['zbfm']:
                if zbfm.value:
                    if zbfm not in repeater2:
                        zbfm.value = isZbfm(zbfm.value, getStr=True)
                    else:
                        zbfm.value = zbfmDelSpace(zbfm.value)

            # 处理合计字段识别错乱， 例： 合计			48.21
            if 'total' in table:
                for cell in sheet[table['total']][0]:
                    cell.value = '合计' if mc.column==cell.column else None

            table['zbfm'].remove(mc)
            table[mcStr],beforeValue = [],''
            for row in sheet[mc.column+str(table['firstZbfm'].row+1)+':'+chr(ord(mc.column)+width)+end]:
                tempStr = ''
                if row[0].value:
                    beforeValue = row[0].value
                else:
                    row[0].value = beforeValue
                for cell in row:
                    if cell.value:
                        tempStr += str(cell.value)
                row[0].value = tempStr
                table[mcStr].append(row[0])

            if mcStr=='zbzm' and width==0:
                table['zbzm'] = self.dealZbzm(table['zbzm'], sheet)

            if 'zbzm' not in table:
                for zbfm in table['zbfm']:
                    if(zbfm.value):
                        if not isZbfm(zbfm.value):
                            if not isObject(zbfm.value):
                                temp = re.findall(u'[\(\（].+[\）\)]', zbfm.value)
                                if temp:
                                    zbfm.value = zbfm.value.replace(temp[0], '')
                                table['zbzm'] = str(zbfm.value)
                                zbfm.value = parseZbzm(zbfm.value)
                                break
                table['zbzm'] = self.getZbzmByTitle(table)

        elif table['type'] == self.TYPE_1_0:
            zbzms = self.findRow(table, sheet, 2, type=isZbzm)
            tempRange = zbzms[1].column+str(zbzms[1].row)+':'+zbzms[-1].column+str(zbzms[-1].row+1)#+1消除偏差
            table['son'],table['object'] = [],[]
            for merged in table['mergedCells']:
                if self.inArea(merged, tempRange)[0]:
                    temp = self.parseArea(merged)
                    table['son'].append({
                        'zbzm': sheet[merged][0][0],
                        'zbfm': sheet[temp[0]+str(temp[1]+1)+':'+temp[2]+str(temp[3]+1)][0]
                    })
            if len(table['son']) == 0:# 说明是单行主指标，单个副指标
                for merged in zbzms[1:]:
                    temp = copy.copy(merged)
                    temp.value = '累计数'
                    table['son'].append({
                        'zbzm': merged,
                        'zbfm': [temp]
                    })
            table['firstZbfm'] = table['son'][0]['zbfm'][0]
            first = table['firstZbfm']
            width = self.getWidth(table)
            beforeValue = ''
            for row in sheet[chr(ord(first.column)-width)+str(table['firstZbfm'].row+1)+':'+chr(ord(first.column)-1)+end]:
                tempStr = ''
                if row[0].value:
                    beforeValue = row[0].value
                else:
                    row[0].value = beforeValue
                for cell in row:
                    if cell.value:
                        tempStr += str(cell.value)
                row[0].value = tempStr
                table['object'].append(row[0])
        else:
            column, table['son'] = 0,[]
            for cell in sheet[table['range']][0]:
                count = 0
                for i in range(tempPaser[1], tempPaser[3]):
                    temp = sheet[cell.column+str(i)]
                    if isZbzm(temp.value):
                        count += 1
                if count >= 2:
                    column = cell.column
                    break
            tempRange = column+str(tempPaser[1])+':'+column+str(tempPaser[3])
            for merged in table['mergedCells']:
                if self.inArea(merged, tempRange)[0]:
                    temp = self.parseArea(merged)
                    table['son'].append({
                        'zbzm': sheet[merged][0][0],
                        'zbfm': [row[0] for row in sheet[chr(ord(temp[0])+1)+str(temp[1])+':'+chr(ord(temp[2])+1)+str(temp[3])]]
                    })
            table['firstZbfm'] = table['son'][0]['zbfm'][0]
            first = table['firstZbfm']
            table['object'] = sheet[chr(ord(first.column)+1)+str(first.row-1)+':'+tempPaser[2]+str(first.row-1)][0]

        if 'object' not in table:
            table['object'] = self.getObjectByTitle(table)
        table['timeCode'] = self.getTimeCodeByTitle(table)
        return table

    def getSingeType(self, table, sheet):
        tempRow,zbzmCount = [],0
        for row in sheet[table['range']]:# 只做了行扫描，列扫描需要再加
            count = 0
            for cell in row:
                if type(cell.value)==type(u''):
                    if isZbzm(cell.value):
                        count += 1
            if count > 1:
                tempRow = row
                break
        for cell in tempRow:
            if cell.value:
                if isZbzm(cell.value):
                    zbzmCount += 1
        table['type'] = self.TYPE_0 if zbzmCount==0 else self.TYPE_1_0
        return table

    def getTableType(self, table, sheet):
        count,temp = 0,[]
        for merged in table['mergedCells']:
            cell = sheet[merged][0][0]
            if cell.value:
                if isZbzm(cell.value):
                    count += 1
                    temp.append(cell)
        if count < 2:
            table = self.getSingeType(table, sheet)
        else:
            col_count, row_count, max_col, max_row = {},{},0,0
            for cell in temp:
                col_count[cell.column] = 1 if cell.column not in col_count else col_count[cell.column]+1
                row_count[cell.row] = 1 if cell.row not in row_count else row_count[cell.row]+1
            for i in col_count:
                if col_count[i]>max_col:
                    max_col = col_count[i]
            for i in row_count:
                if row_count[i]>max_row:
                    max_row = row_count[i]
            table['type'] = self.TYPE_1_0 if max_row>max_col else self.TYPE_1_1
        return table

    def getBase(self, sheet):
        mergedCells = sheet._merged_cells
        tables = {}
        for merged in mergedCells:
            inArea = False
            mergedParse = self.parseArea(merged)
            for t in tables:
                inArea, bigStr = self.inArea(merged, t)
                if inArea:
                    tables[bigStr]['mergedCells'].append(merged)
                    tempValue = sheet[merged][0][0].value
                    if tempValue:
                        if re.findall('合.*计', str(tempValue)) and not self.isEnd(tempValue):
                            tables[bigStr]['total'] = merged
                    break
            if not inArea and not self.isEnd(sheet[merged][0][0].value, mList=ignoreRows) and mergedParse[0]!=mergedParse[2]:
                i,rowNone,temp = 1,False,self.parseArea(merged)
                while(not rowNone):
                    rowNone = True if i>4 else False
                    cells = sheet[temp[0]+str(temp[3]+i)+':'+temp[2]+str(temp[3]+i)][0]
                    for cell in cells:
                        if cell.value:
                            if self.isEnd(cell.value):
                                i += 1
                                break
                            if self.similar(cell.value, sheet[merged][0][0].value):
                                break
                            rowNone = False
                            break
                    i += 1
                if rowNone:
                    i -= 1
                table = self.parseTitle(merged, sheet)
                if table:
                    rangeStr = temp[0]+str(temp[1])+':'+temp[2]+str(temp[3]+i-1)
                    table['range'] = rangeStr
                    tables[rangeStr] = table
        return tables

    def writeExcel(self, sheet, wb):
        if sheet.sheet_state == 'veryHidden':
            return
        print sheet.title,'==================================================================================================================='
        ws = wb.create_sheet()
        ws.title = re.sub('[\d\-\.\、]+','',sheet.title)
        ws.append(['时间', '地区', '指标主名', '指标副名', '指标值', '表名'])
        tables = self.getBase(sheet)
        for t in tables:
            tables[t] = self.getTableType(tables[t], sheet)
            tables[t] = self.getZbObject(tables[t], sheet)
            t = tables[t]
            if t['type'] == self.TYPE_0:
                objStr = 'zbzm' if type(t['object']) == type('') or type(t['object']) == type(u'') else 'object'
                for obj in t[objStr]:
                    if obj.value:
                        tempObj = t['object'] if objStr == 'zbzm' else obj.value
                        tempZbzm = obj.value if objStr == 'zbzm' else t['zbzm']
                        for zbfm in t['zbfm']:
                            if zbfm.value:
                                if not self.isIgnoreZbfm(isZbfm(zbfm.value, getStr=True)):
                                    value = sheet[zbfm.column+str(obj.row)].value if sheet[zbfm.column+str(obj.row)].value else ''
                                    ws.append([t['timeCode'], delPreStr(tempObj), delPreStr(tempZbzm), zbfm.value, value, delTime(delPreStr(t['title']))])
            else:
                for son in t['son']:
                    for obj in t['object']:
                        if obj.value:
                            for zbfm in son['zbfm']:
                                if zbfm.value:
                                    if not self.isIgnoreZbfm(isZbfm(zbfm.value, getStr=True)):
                                        value = sheet[zbfm.column + str(obj.row)].value if sheet[zbfm.column + str(obj.row)].value else ''
                                        # print value, zbfm.column, str(obj.row), t['title'], son['zbzm'].value, zbfm.value, obj.value
                                        if t['type'] == self.TYPE_1_1:
                                            value = sheet[obj.column + str(zbfm.row)].value if sheet[obj.column + str(zbfm.row)].value else ''
                                        ws.append([t['timeCode'], delPreStr(obj.value), delPreStr(isZbzm(son['zbzm'].value, getStr=True)), isZbfm(zbfm.value, getStr=True), value, isZbzm(son['zbzm'].value, getStr=True)])
                                        # ws.append([t['timeCode'], delPreStr(obj.value), delPreStr(isZbzm(son['zbzm'].value, getStr=True)), isZbfm(zbfm.value, getStr=True), value, isZbzm(son['zbzm'].value, getStr=True), value, zbfm.column, str(obj.row), t['title'], son['zbzm'].value, zbfm.value, obj.value])


    def inputDBTemp(self, sheet, areaName, excelName):
        if sheet.sheet_state == 'veryHidden':
            return
        print sheet.title,'==================================================================================================================='
        ws = []
        sheet_name = re.sub('[\d\-\.\、]+','',sheet.title)
        tables = self.getBase(sheet)
        for t in tables:
            tables[t] = self.getTableType(tables[t], sheet)
            tables[t] = self.getZbObject(tables[t], sheet)
            t = tables[t]
            if t['type'] == self.TYPE_0:
                objStr = 'zbzm' if type(t['object']) == type('') or type(t['object']) == type(u'') else 'object'
                for obj in t[objStr]:
                    if obj.value:
                        tempObj = t['object'] if objStr == 'zbzm' else obj.value
                        tempZbzm = obj.value if objStr == 'zbzm' else t['zbzm']
                        for zbfm in t['zbfm']:
                            if zbfm.value:
                                if not self.isIgnoreZbfm(isZbfm(zbfm.value, getStr=True)):
                                    value = sheet[zbfm.column+str(obj.row)].value if sheet[zbfm.column+str(obj.row)].value else ''
                                    ws.append([t['timeCode'], delPreStr(tempObj), delPreStr(tempZbzm), zbfm.value, value, delTime(delPreStr(t['title'])), sheet_name])
            else:
                for son in t['son']:
                    for obj in t['object']:
                        if obj.value:
                            for zbfm in son['zbfm']:
                                if zbfm.value:
                                    if not self.isIgnoreZbfm(isZbfm(zbfm.value, getStr=True)):
                                        value = sheet[zbfm.column + str(obj.row)].value if sheet[zbfm.column + str(obj.row)].value else ''
                                        if t['type'] == self.TYPE_1_1:
                                            value = sheet[obj.column + str(zbfm.row)].value if sheet[obj.column + str(zbfm.row)].value else ''
                                        ws.append([t['timeCode'], delPreStr(obj.value), delPreStr(isZbzm(son['zbzm'].value, getStr=True)), isZbfm(zbfm.value, getStr=True), value, isZbzm(son['zbzm'].value, getStr=True), sheet_name])

        session = DBSession
        for row in ws:
            sqlText = "insert into dw_zhs_f_month_imp(time_name,object_name,zbzm_name,zbfm_name,indication_value_str,table_name,sheet_name,area_name, excel_name) values('"+str(row[0])+"','"+str(row[1])+"','"+str(row[2])+"','"+str(row[3])+"','"\
                      +str(row[4])+"','"+str(row[5])+"','"+str(row[6])+"','"+areaName+"', '"+excelName+"')"
            session.execute(sqlText)
        CONN.commit()


    def mPrint(self, sheet):
        if sheet.sheet_state == 'veryHidden':
            return
        tables = self.getBase(sheet)
        for t in tables:
            tables[t] = self.getTableType(tables[t], sheet)
            tables[t] = self.getZbObject(tables[t], sheet)
            for i in tables[t]:
                if i=='zbfm' or i=='object' or i=='zbzm':
                    print i, ':',
                    if type(tables[t][i]) == type('') or type(tables[t][i]) == type(u''):
                        print tables[t][i],'---------'
                    else:
                        for j in tables[t][i]:
                            print j.value.replace('\n', '').replace(' ', '') if j.value else j.value,
                        print ''
                elif i == 'son':
                    for j in tables[t][i]:
                        print '主指标 :',j['zbzm'].value.replace('\n', '').replace(' ', '') if j['zbzm'].value else j['zbzm'].value,'\n副指标 :',
                        for k in j['zbfm']:
                            print str(k.value).replace('\n', '').replace(' ', '') if k.value else k.value,
                        print '\n+++++++++++++++++++++++++++++++++++++++'
                else:
                    print i,':', tables[t][i]
            print '----------------------------------------------------------------'

    def close(self):
        self.wb.close()

def getExcel():
    e = MExcel('../../file/8.xlsx')
    wb = Workbook()
    for sheetName in e.wb.get_sheet_names():
        # e.mPrint(e.wb.get_sheet_by_name(sheetName))
        e.writeExcel(e.wb.get_sheet_by_name(sheetName), wb)
    path = 'D:test9.xlsx'
    wb.save(path)
    wb = load_workbook(path)
    wb.remove_sheet(wb.active)
    wb.save(path)

def mPrint():
    e = MExcel('../../file/8.xlsx')
    for sheetName in e.wb.get_sheet_names():
        e.mPrint(e.wb.get_sheet_by_name(sheetName))

def inputDB(path, areaName, excelName):
    e = MExcel(path)
    for sheetName in e.wb.get_sheet_names():
        e.inputDBTemp(e.wb.get_sheet_by_name(sheetName), areaName, excelName)
        # e.inputDBTemp(e.wb.get_sheet_by_name(sheetName), '龙华统计局','统计月报')


if __name__ == '__main__':
    getExcel()
    # inputDB('../../file/4.xlsx','宝安统计局','统计月报')
    # inputDB('../../file/8.xlsx','龙华统计局','统计月报')
    # mPrint()





